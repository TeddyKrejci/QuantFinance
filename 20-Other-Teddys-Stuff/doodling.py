
from firstPack.main_teddyPack import main_report
from firstPack.subPack.sub_teddyPack import sub_report
import numpy as np
from colorama import Fore
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from first_module import teds_func
#from firstPack import main_teddyPack






# first command
print("Hi")


# some initial priting to see if python runs
for i in range(1, 4):
    print(i)

print(np.cos(3*np.pi))
#print(Fore.GREEN + "Here's some green text.")
#print(Fore. WHITE + "now back to white..")


# testing work with xl workbooks

# create and save a workbook
mybook = Workbook()
file_name = 'mysecond_xl.xlsx'
mysheet = mybook.active
mysheet.title = 'great_sheet_name'
for row in range(1, 40):
    mysheet.append(range(600))

othersheet=mybook.create_sheet(title='better_name')    

othersheet['B2']=554

mybook.save(filename = file_name)

# load an existing workbook

ext_book=load_workbook(filename = 'EIMI_holdings.xlsx')

sheet_to_print=ext_book['Pivots']

print(sheet_to_print['D18'].value)


teds_func("Tower of Power")

# call function from user defined package

main_report()

sub_report()